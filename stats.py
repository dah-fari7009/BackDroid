#!/usr/bin/python

import os
import sys
import subprocess
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd


apk_dir = sys.argv[1]
excel_name =  sys.argv[2] if len(sys.argv) > 2 else 'analysis'
log_dir = sys.argv[3] if len(sys.argv) > 3 else apk_dir
#log_file = sys.argv[1] if len(sys.argv) > 1 else 'data.log'
#apk_path = sys.argv[2] if len(sys.argv) > 2 else 'SmartSwitch.apk'
#out_dir = apk_dir

#log_file_id = "LOG"
dyn_log_dir = "/data/faridah/experiments/all_logs"
log_file_id = "DEBUG_INTERMEDIATE_LOG"
ape_id="APE"
manual_id="MANUAL"
file_ext="log"

excel_file = None


#output = {'Activity': [], '#Stmts': [], 'Stmts': [], '#Callers': [], 'Caller entrypoints': []}

#TODO handle activity-aliases; <activity-alias name="" targetActivity="">
# alias can be reported at runtime (in reached activities), should be counted as if targetActivity was reached
#look more into it

def write_results(data, sheetname, summary=False):
    content = pd.DataFrame(data)
    if not summary:
        content = content.sort_values(by='Activity', key=lambda col: col.str.lower())
    num_columns = content.shape[1]
    cell_width = 200/num_columns
    writer = None
    if os.path.isfile(excel_file):
        #file already exists, we append
        #print("File already exists, appending")
        writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists="replace")
        writer.book = load_workbook(excel_file)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets) 
        
        content.to_excel(writer, sheet_name=sheetname, index=False)
        ws = writer.sheets[sheetname]
        #Settings (width of columns, text wrap)
        for col in range(1, num_columns + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = cell_width 
            #print(writer.sheets[sheetname])
            #writer.sheets[sheetname][letter].alignment = Alignment(wrap_text=True)
    else:#
        writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')

        content.to_excel(writer, sheet_name=sheetname, index=False)

        #Settings (width of columns, text wrap)
        format = writer.book.add_format({'text_wrap': True})
        writer.sheets[sheetname].set_column(0, num_columns, cell_width, cell_format=format)
        #writer.sheets[sheetname].conditional_format('', {
        #    'type': 'formula',
        #    'criteria':'',
        #    'value': 'N/A',
        #    'format': custom_format
        #})
    writer.save()
    #writer.close()


#Get package names from all apks in folder
def get_package_names(apk_path):
    package_names = {}
    #print(apk_dir)
    if not os.path.isdir(apk_path) and apk_path.endswith('.apk'):
        print(f"=======================================Parsing apk {apk_path}")
        match = None
        cmd = ["sh", "./scripts/package.sh", f"{apk_path}"]
        p = subprocess.Popen(cmd, stdout=subprocess.PIPE)  
        for line in p.stdout:
            match = line.decode('utf8').replace("\n","").replace("'","")#.split(' ')[1]
            #match = match.split("=")[1].replace("'", "")
            #print(match)
        p.wait()

        if match is not None:
            return (match, get_comp_from_manifest(apk_path,"activity"), get_comp_from_manifest(apk_path,"service"), get_comp_from_manifest(apk_path,"receiver"))
    return None

def get_comp_from_manifest(apk_path, type="activity"):
    cmd2 = ["sh", "./scripts/activity-list.sh", f"{apk_path}", type]
    package_names = []
    p2 = subprocess.Popen(cmd2, stdout=subprocess.PIPE) 
    for line in p2.stdout:
        package_names.append(line.decode('utf8').replace("\n","").replace("'",""))
    p2.wait()
    #print(f"Found {type} for {apk_path} {package_names}")
    return package_names

def build_act(full_name):
    splits = full_name.replace("\n","").replace("'","").replace("}","").split("/")
    #print(f"Splits: {splits}")
    if(len(splits) < 2):
        return "N/A"
    return f"{splits[0]}{splits[1]}" if splits[1].startswith(".") else splits[1]

def get_implicit_intents_from_manifest(apk_path, activities):
    #run script to get the activities and the filters?
    cmd = ["sh", "./scripts/manifest.sh", f"{apk_path}"]
    p = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    
    intents_map = {}
    activity = None
    action = None
    #todo issue in script, see Audible
    for line in p.stdout:
        if b"E: activity (" in line:
            activity = "N/A"
            action = None
        if b"E: action (" in line:
            action = "N/A"
        elif b"A: android:name(" in line: #android:name(0x01010003)="com.sec.android.easyMover.ui.IntroduceSamsungActivity"
            if activity == "N/A": #seen activity 
                match = line.decode('utf-8').split("\"")[1]
                if match in activities:
                    activity = match
            elif not activity is None: #activity is set and we see name
                if action == "N/A":
                    match = line.decode('utf-8').split("\"")[1]
                    action = match
                    if activity in intents_map:#action in intents_map:
                        intents_map[activity].append(action)
                        #intents_map[action].append(activity)
                    else:
                        intents_map[activity] = [action]
                        #intents_map[action] = [activity]
                    
                else: #saw another name, but no action
                    activity = None
                    action = None
            
    p.wait()
    return intents_map

def get_implicit_full(apk_path, activities):
    implicit_activities = get_implicit_intents_from_manifest(apk_path, activities)
    implicit_full = set()
    for key,val in implicit_activities.items():
        implicit_full.update(val)
    return implicit_full


def get_reached_activities(package_name):
    reached_activities = set()
    key = package_name
    ape_log = f"{dyn_log_dir}/{ape_id}_{key}.{file_ext}"
    manual_log = f"{dyn_log_dir}/{manual_id}_{key}.{file_ext}"
    valid_log = f"{dyn_log_dir}/VALIDATION_{key}.{file_ext}"
    #valid_log = f"{dyn_log_dir}/VALIDATION_{key}.{file_ext}"

    #print(f"Looking for {ape_log} and {manual_log}")
    #for each line in the file, mark activity as reached by ape, package[key][line] = reached
    ape_activities = None
    manual_activities = None

    if(os.path.isfile(ape_log)):
        #print(f"Found ape file {ape_log}")
        lines = open(ape_log).readlines()
        if len(lines) > 0:
            ape_activities = set(map(lambda x: build_act(x) , lines)) #todo except last line
        #print(ape_activities)
    if(os.path.isfile(manual_log)):
        #print(f"Found manual file {manual_log}")
        lines = open(manual_log).readlines()
        if len(lines) > 0:
            manual_activities = set(map(lambda x: build_act(x) , lines))
        #print(manual_activities)
    if not ape_activities is None:
        reached_activities.update(ape_activities)
    if not manual_activities is None:
        reached_activities.update(manual_activities)
    reached_activities.remove('N/A')
    return reached_activities


def breakdown_for_notfound(apk_name, found_activities, reached_activities, implicit_activities, all_activities):
    found_activities = [act for act in found_activities if act in all_activities]
    if len(found_activities) == len(all_activities):
        return "N/A"
    unreached_tot = len(all_activities) - len(reached_activities)

    #todo deal with activity aliases
    total = len(all_activities)
    resolved = len(found_activities)
    #implicit = len(implicit_activities)
    #print(f"Out of {total} activities, {unreached_tot} unreached, {resolved}/{total} explicit resolved, {len(implicit_activities.keys())}/{total} implicit")
    
    notfound_activities = [act for act in all_activities if (act not in reached_activities and act not in found_activities )] #all activities not found that were not reached (don't care about reached)
    num = len(notfound_activities)
    print(f"{num}/{unreached_tot} not found activities by the tool")

    new_set = set()
    #new_set.update(implicit_activities)
    new_set.update(found_activities)
    #implicit_full.update(data.keys())
    (alternative, tool_limitations) = ([], [])
    if num > 0:
        #search for class definition in file

        # todo, search for implicit intents and if not found, then alternative
        for act in notfound_activities:
            applog = f"{log_dir}/{apk_name}_dexdump.log"
            formatted_name = bytecode_format(act)
    
            # Search "const-class .*, Lcom/lge/app1/fota/HttpServerService;"
            cmd = f"cat {applog} | grep -e \"const-class .*, {formatted_name}\""
            #print(f"Executing {cmd}")
            process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            (out, err) = process.communicate()  #TODO error handling
            if out.decode() != '':                        
                tool_limitations.append(act)
                #print(out)  
            else: # couldn't find class name
                #try the intent
                if act in implicit_activities:
                    actions = '|'.join(implicit_activities[act])
                    cmd = f'cat {applog} | grep -E "const-string.*, \\"({actions})\\""'
                    print(f"Executing {cmd} for {act}")
                    process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    (out, err) = process.communicate()  #TODO error handling
                    if out.decode() != '':
                       new_set.add(act)
                    else:
                        #print(f"Couldn't find {act}")
                        alternative.append(act)
                else:
                    alternative.append(act)
        #print(f"Potential alt: {alternative}")
        max_potential = [act for act in new_set if act not in reached_activities]
        print(f"Max unreached resolvable: {len(max_potential)}/{unreached_tot}")
        print(f"Breakdown of {unreached_tot} (#pot: {len(max_potential)}, #alt: {len(alternative)}, #lim: {len(tool_limitations)})")
    return "DONE"



def bytecode_format(act):
    b_string = "L"+act.replace(".","/")+";"
    return b_string

#Info for <com.sec.android.easyMover.ui.MainActivity: void runMenuItem(com.sec.android.easyMover.host.ActivityBase$UiMenuType)>
#Targets:  [class "com/sec/android/easyMover/ui/SendOrReceiveActivity"] or ["com/sec/android/easyMover/ui/SendOrReceiveActivity"]
#Tail nodes: [BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity]]
#Fake tail nodes [BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity$28: void onStateReceived(com.samsung.android.sdk.bixby.data.State)>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity$28], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity$28: void onStateReceived(com.samsung.android.sdk.bixby.data.State)>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity$28], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void <init>()>, unit=nop]]
#Field nodes: [BDGUnit [msig=<com.sec.android.easyMover.IAConstants$SelectedExType: void <clinit>()>, unit=$r0 = new com.sec.android.easyMover.IAConstants$SelectedExType]]
def parse_output(log_file, reached_activities, activities, services, receivers):
    data = {}
    
    cmd = ['grep', '-e', "Exception in thread \"main\"", log_file]
    #print(f"Executing {cmd}")

    p = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    (out, err) = p.communicate()
    potential_exception = out.decode()
    if potential_exception != '':
        print(f"{log_file}: Exception {potential_exception}")
        return (None, potential_exception)
    cmd = ['grep', '-e', "Info for", "-e", "Targets", "-e", "Tail nodes", "-e", "Fake tail nodes", log_file]
    #print(f"Executing {cmd}")
    p = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    
    stmt = None
    targets = []
    callers = None
    for line in p.stdout:
        if b"Info for " in line:
            match = (line.decode('utf-8').split('<')[1])[:-2] #com.sec.android.easyMover.ui.MainActivity: void runMenuItem(com.sec.android.easyMover.host.ActivityBase$UiMenuType)>
            #match = match.split("=")[1].replace("'", "")
            stmt = match
        elif b"Targets: " in line:
            #print(line.decode("utf-8").replace("\n",""))
            match = (line.decode('utf-8').split("Targets: ")[1][1:-1].split(', ')) #[class "com/sec/android/easyMover/ui/SendOrReceiveActivity"] or ["com/sec/android/easyMover/ui/SendOrReceiveActivity"]
            targets = [val.replace("class","").replace("\"","").replace("/",".").replace("]","").strip() for val in match if val.startswith("class") or val.startswith("\"")]
            #print(f"Intermediate targets {targets}")
            to_resolve = [val for val in targets if val not in activities]
            #if (len(to_resolve) > 0):
                #print(f"Extra targets to resolve: {len(to_resolve)}")
            targets = [val for val in targets if val in activities]
            #print(f"Obtained targets {targets}\n")
        elif b"Tail nodes" in line: #[BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity], BDGUnit [msig=<com.sec.android.easyMover.ui.MainActivity: void onResume()>, unit=r0 := @this: com.sec.android.easyMover.ui.MainActivity]]
            match = (line.decode('utf-8')).split("BDGUnit [msig=<")[1:]
            callers = [val.split(">, unit=")[0] for val in match]
        elif b"Fake tail nodes" in line:
            if callers is None:
                match = (line.decode('utf-8')).split("BDGUnit [msig=<")[1:]
                callers = [val.split(">, unit=")[0] for val in match]
            for target in targets:
                if target in activities:
                    e_stmts, e_callers = data.get(target,[[], []])
                    data[target] = [e_stmts + [stmt], e_callers + callers]
            callers = None
            stmt = None
            targets = None
    p.wait()
    #print(f"Parsed info {data.keys()}")

    output = {'Activity': [], '#Stmts': [], 'Stmts': [], 'Has caller activity?': [], 'Has caller component?': [],'#Callers': [], 'Callers': [], 'Is not reached dynamically?': [], 'Is not reached AND has caller component': [], 'Is not reached AND has caller': []}

    index = 2 # first lines
    max_val = len(data.keys()) + index
    output['Activity'].append(f"=COUNTIF(A3:A{max_val},\"*\")")
    output['Activity'].extend(data.keys())


    output['#Stmts'].append(f"=SUM(B3:B{max_val})")
    output['Stmts'].append("")
    output['Has caller activity?'].append(f"=COUNTIF(D3:D{max_val},\"<>NO\")")
    output['Has caller component?'].append(f"=COUNTIF(E3:E{max_val},\"<>NO\")")
    output['#Callers'].append("")
    output['Callers'].append("")
    output['Is not reached dynamically?'].append(f"=COUNTIF(H3:H{max_val},\"YES\")")
    output['Is not reached AND has caller component'].append(f"=COUNTIF(I3:I{max_val}, \"YES\")")
    output['Is not reached AND has caller'].append(f"=COUNTIF(J3:J{max_val}, \"YES\")")
    for act, info in data.items():
        num_stmts = len(info[0])
        stmts = "\n".join(info[0])
        callers = list(set(info[1]))
        num_callers = len(callers)
        has_caller_activity = check_caller_component(act, callers, activities)
        has_caller_component = has_caller_activity or (check_caller_component(act, callers, services) or check_caller_component(act, callers, receivers))
        has_caller = num_callers > 0
        is_not_reached_dynamically = "NO" if act in reached_activities else "YES"
        if len(reached_activities) == 0:
            is_not_reached_dynamically = "UNK."
        callers = "\n".join(callers)
        output['#Stmts'].append(num_stmts)
        output['Stmts'].append(stmts)
        output['Has caller activity?'].append("YES" if has_caller_activity else "NO")
        output['Has caller component?'].append("YES" if has_caller_component else "NO")
        output['#Callers'].append(num_callers)
        output['Callers'].append(callers)
        output['Is not reached dynamically?'].append(is_not_reached_dynamically)
        output['Is not reached AND has caller component'].append("YES" if is_not_reached_dynamically == "YES" and has_caller_component else "NO")
        output['Is not reached AND has caller'].append("YES" if is_not_reached_dynamically == "YES" and has_caller else "NO")

    return (output, None)


def check_caller_component(act, callers, components):
    if len(callers) == 0:
        return False
    unique_callers = [format(val) for val in callers if format_val(val) is not act and format_val(val) in components]
    return True if len(unique_callers) > 0 else False

def format_val(val):
    #print(val.split(":")[0])
    return val.split(":")[0]
#def analysis_summary():
    #Build the summary table
    #Apk Package #Activities (total) #Activities (resolved ICC) #Activities (found caller act)

def prep_summary(max):
    summary = {'Apk': [], 'Package Name': [], '#Activities (total)': [],  '#Activities (unreached)': [], '#Activities (resolved ICC)': [], '%Activities (resolved ICC)': [], '#Unreached activities (resolved ICC)': [], '%Unreached activities (resolved ICC)': [], '#Activities (with caller activity)': [], '%Activities (with caller activity)': [], '#Activities (with caller comp)': [],'%Activities (with caller comp)': [], '#Unreached activities (with caller comp)': [], '%Unreached activities (with caller comp)': []}
    summary["Apk"].append("")
    summary["Package Name"].append("")
    summary["#Activities (total)"].append("")
    summary["#Activities (unreached)"].append("")

    summary["#Activities (resolved ICC)"].append("")
    summary["#Unreached activities (resolved ICC)"].append("")
    summary["#Activities (with caller activity)"].append("")
    summary["#Activities (with caller comp)"].append("")
    summary["#Unreached activities (with caller comp)"].append("")
    
    summary["%Activities (resolved ICC)"].append(f"=AVERAGEA(F3:F{max})")
    summary["%Unreached activities (resolved ICC)"].append(f"=AVERAGEA(H3:H{max})")
    summary["%Activities (with caller activity)"].append(f"=AVERAGEA(J3:J{max})")
    summary["%Activities (with caller comp)"].append(f"=AVERAGEA(L3:L{max})")
    summary["%Unreached activities (with caller comp)"].append(f"=AVERAGEA(N3:N{max})")
    
    return summary

def main():
    global excel_file
    global excel_name

    excel_file = f"{excel_name}.xlsx" 
    print("Collecting stats from BackDroid")


    summary = prep_summary(50) #to automate
    summary_sheet_name = "Summary"
    #write_results(summary, summary_sheet_name, summary=True)
    index = 2
    for apk in os.listdir(apk_dir):
        if apk.endswith('.apk'):
            apk_path = f"{apk_dir}/{apk}"
            apk_name = apk[:-4]
            sheet_name = apk_name
            #Get info about the apk
            package_name, activities, services, receivers = get_package_names(apk_path)
            num_activities = len(activities)
            #print(f"Found apk {apk_path} => {package_name}")
            log_file = f"{log_dir}/{log_file_id}_{package_name}.log"
            if(os.path.exists(log_file)):
                index = index + 1
                #print(f"Found log file: {log_file}")
                reached_activities = get_reached_activities(package_name)
                num_reached_activities = len(reached_activities)
                implicit_full = get_implicit_intents_from_manifest(apk_path, activities) #get_implicit_full(apk_path, activities)
                #print(f"Found {num_reached_activities} reached activities for {apk_path}")
                (data, exc) = parse_output(log_file, reached_activities, activities, services, receivers)
                summary["Apk"].append(apk_name)
                summary["Package Name"].append(package_name)
                summary["#Activities (total)"].append(num_activities)
                summary["#Activities (unreached)"].append((num_activities-num_reached_activities))
                if exc is None:
                    info = breakdown_for_notfound(apk_name, data['Activity'], reached_activities, implicit_full, activities)
                    summary["#Activities (resolved ICC)"].append(f"='{sheet_name}'!A2")
                    summary["%Activities (resolved ICC)"].append(f"=(100 * E{index}/C{index})")
                    summary["#Unreached activities (resolved ICC)"].append(f"='{sheet_name}'!H2")
                    summary["%Unreached activities (resolved ICC)"].append(f"=(100 * G{index}/D{index} )")
                    summary["#Activities (with caller activity)"].append(f"='{sheet_name}'!D2")
                    summary["%Activities (with caller activity)"].append(f"=(100 * I{index}/C{index})")
                    summary["#Activities (with caller comp)"].append(f"='{sheet_name}'!E2")
                    summary["%Activities (with caller comp)"].append(f"=(100 * K{index}/C{index})")
                    summary["#Unreached activities (with caller comp)"].append(f"='{sheet_name}'!I2")
                    summary["%Unreached activities (with caller comp)"].append(f"=(100 * M{index}/D{index})")
                    #write_results(data, sheet_name)
                else:
                    summary["#Activities (resolved ICC)"].append(exc)
                    summary["%Activities (resolved ICC)"].append("")
                    summary["#Unreached activities (resolved ICC)"].append("")
                    summary["%Unreached activities (resolved ICC)"].append("")
                    summary["#Activities (with caller activity)"].append("")
                    summary["%Activities (with caller activity)"].append("")
                    summary["#Activities (with caller comp)"].append("")
                    summary["%Activities (with caller comp)"].append("")
                    summary["#Unreached activities (with caller comp)"].append("")
                    summary["%Unreached activities (with caller comp)"].append("")
    #write_results(summary, summary_sheet_name, summary=True)

if __name__ == '__main__':
    main()



# Landroid/content/Context;.startActivity:(Landroid/content/Intent;)V  
#    # grep dynamic code loading
#    cmd = 'cat %s | grep -e "Ldalvik/system/DexClassLoader;.loadClass:(Ljava/lang/String;)Ljava/lang/Class;" -e "Ldalvik/system/PathClassLoader;.loadClass:(Ljava/lang/String;Z)Ljava/lang/Class;"' % applog 
#    process = Popen(cmd, shell=True, stdout=PIPE, stderr=PIPE)
#    (out, err) = process.communicate()  #TODO error handling
#    if out.decode() != '':
#        resdcl = 1